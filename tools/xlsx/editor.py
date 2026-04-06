"""XLSX editing — update cells, add sheets, modify data."""

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill


def update_cell(path: str, sheet_name: str, cell_ref: str, value, output_path: str) -> str:
    """Update a single cell value. cell_ref is like 'A1', 'B5'."""
    wb = load_workbook(path)
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
    ws[cell_ref] = value
    wb.save(output_path)
    return output_path


def update_cells(path: str, sheet_name: str, updates: dict, output_path: str) -> str:
    """Update multiple cells. updates is {cell_ref: value}."""
    wb = load_workbook(path)
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
    for cell_ref, value in updates.items():
        ws[cell_ref] = value
    wb.save(output_path)
    return output_path


def add_sheet(path: str, sheet_name: str, headers: list[str] | None, output_path: str) -> str:
    """Add a new sheet to an existing workbook."""
    wb = load_workbook(path)
    ws = wb.create_sheet(title=sheet_name[:31])
    if headers:
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
    wb.save(output_path)
    return output_path


def append_row(path: str, sheet_name: str, values: list, output_path: str) -> str:
    """Append a row to an existing sheet."""
    wb = load_workbook(path)
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
    ws.append(values)
    wb.save(output_path)
    return output_path


def delete_sheet(path: str, sheet_name: str, output_path: str) -> str:
    """Delete a sheet from a workbook."""
    wb = load_workbook(path)
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    wb.save(output_path)
    return output_path

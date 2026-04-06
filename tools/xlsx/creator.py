"""XLSX creation from structured data."""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def create_from_data(data: dict, output_path: str) -> str:
    """Create an XLSX from dict data.

    data format: {sheet_name: [{"col": val, ...}, ...]}
    """
    wb = Workbook()
    wb.remove(wb.active)

    for sheet_name, rows in data.items():
        ws = wb.create_sheet(title=sheet_name[:31])  # Excel 31-char limit

        if not rows:
            continue

        headers = list(rows[0].keys())

        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center")
            cell.border = THIN_BORDER

        # Write data
        for row_idx, record in enumerate(rows, 2):
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row_idx, column=col, value=record.get(header, ""))
                cell.border = THIN_BORDER

        # Auto-width columns
        for col in range(1, len(headers) + 1):
            max_len = max(
                len(str(ws.cell(row=r, column=col).value or ""))
                for r in range(1, len(rows) + 2)
            )
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = min(max_len + 3, 50)

    wb.save(output_path)
    return output_path


def create_simple(headers: list[str], rows: list[list], output_path: str,
                  sheet_name: str = "Sheet1") -> str:
    """Create a simple XLSX from headers and row data."""
    data = {}
    records = []
    for row in rows:
        record = {}
        for i, header in enumerate(headers):
            record[header] = row[i] if i < len(row) else ""
        records.append(record)
    data[sheet_name] = records
    return create_from_data(data, output_path)

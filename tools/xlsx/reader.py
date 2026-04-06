"""XLSX reading and data extraction."""

from openpyxl import load_workbook


def extract_data(path: str, sheet_name: str | None = None) -> dict:
    """Extract all data from an XLSX file as {sheet_name: [{col: val, ...}]}.

    If sheet_name is given, only that sheet is extracted.
    """
    wb = load_workbook(path, data_only=True)
    sheets = [sheet_name] if sheet_name else wb.sheetnames
    result = {}

    for name in sheets:
        if name not in wb.sheetnames:
            continue
        ws = wb[name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            result[name] = []
            continue

        headers = [str(h) if h else f"col_{i}" for i, h in enumerate(rows[0])]
        data = []
        for row in rows[1:]:
            record = {}
            for i, val in enumerate(row):
                if i < len(headers):
                    record[headers[i]] = val
            data.append(record)
        result[name] = data

    return result


def extract_text(path: str, sheet_name: str | None = None) -> str:
    """Extract all cell values as plain text."""
    wb = load_workbook(path, data_only=True)
    sheets = [sheet_name] if sheet_name else wb.sheetnames
    lines = []

    for name in sheets:
        if name not in wb.sheetnames:
            continue
        ws = wb[name]
        lines.append(f"=== Sheet: {name} ===")
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            lines.append("\t".join(cells))
        lines.append("")

    return "\n".join(lines)


def get_sheet_names(path: str) -> list[str]:
    """Get all sheet names in an XLSX file."""
    wb = load_workbook(path, read_only=True)
    return wb.sheetnames
